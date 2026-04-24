import flet as ft
from datetime import datetime
from utils.helpers import parse_date, converter_tempo_para_segundos, formatar_segundos_para_tempo, obter_duracao_efetiva, format_date_br, check_status_match
import asyncio
import os
import tkinter as tk
from tkinter import filedialog

# Imports pesados no topo
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    pass

class RelatorioView(ft.Column):
    def __init__(self, controller, on_back=None):
        super().__init__(expand=True, spacing=20, scroll=ft.ScrollMode.AUTO)
        self.controller = controller
        self.on_back = on_back
        self.filtro_periodo = "Semestral"
        self.filtro_ano = None
        self.filtro_tipo = "Todos"
        self.filtro_status = "Todos"
        self.filtro_modalidade = "Todos"
        self.filtro_local = "Todos"
        self.filtro_operador = "Todos"
        self.termo_busca = ""
        
        self.tabela_resumo = ft.DataTable(
            columns=[], rows=[], border=ft.border.all(1, ft.Colors.WHITE10), border_radius=10,
            heading_row_color=ft.Colors.with_opacity(0.15, ft.Colors.BLUE_400), heading_row_height=50,
            data_row_min_height=40, data_row_max_height=50, column_spacing=20,
        )
        self.tabela_detalhada = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Data")),
                ft.DataColumn(ft.Text("Horário")),
                ft.DataColumn(ft.Text("Evento")),
                ft.DataColumn(ft.Text("Responsável")),
                ft.DataColumn(ft.Text("Tipo")),
                ft.DataColumn(ft.Text("Modalidade")),
                ft.DataColumn(ft.Text("Local")),
                ft.DataColumn(ft.Text("Operador")),
                ft.DataColumn(ft.Text("Público"), numeric=True),
                ft.DataColumn(ft.Text("Duração")),
                ft.DataColumn(ft.Text("Status")),
            ],
            rows=[], border=ft.border.all(1, ft.Colors.WHITE10), border_radius=10,
            heading_row_color=ft.Colors.with_opacity(0.15, ft.Colors.ORANGE_400), heading_row_height=50,
            column_spacing=15, horizontal_margin=10,
        )
        self.init_ui()

    def init_ui(self):
        # Barra de Busca
        self.txt_busca = ft.TextField(
            label="Pesquisar no relatório...",
            prefix_icon=ft.Icons.SEARCH,
            on_change=self.on_filtro_change,
            expand=True,
            height=45,
            text_size=13
        )

        # Filtros de Agrupamento
        self.dd_periodo = ft.Dropdown(label="Agrupamento", options=[ft.dropdown.Option("Semestral"), ft.dropdown.Option("Anual"), ft.dropdown.Option("Mensal")], value=self.filtro_periodo, on_select=self.on_filtro_change, border=ft.InputBorder.NONE, width=150)
        
        anos = self.get_anos_disponiveis()
        self.dd_ano = ft.Dropdown(label="Ano", options=[ft.dropdown.Option("Todos")] + [ft.dropdown.Option(str(a)) for a in anos], value="Todos", on_select=self.on_filtro_change, border=ft.InputBorder.NONE, width=110)
        
        tipos = self.get_tipos_disponiveis()
        self.dd_tipo = ft.Dropdown(label="Tipo", options=[ft.dropdown.Option("Todos")] + [ft.dropdown.Option(t) for t in tipos], value="Todos", on_select=self.on_filtro_change, border=ft.InputBorder.NONE, width=150)

        # Novos Filtros
        self.dd_status = ft.Dropdown(
            label="Status", 
            options=[ft.dropdown.Option("Todos"), ft.dropdown.Option("Agendado"), ft.dropdown.Option("Finalizado"), ft.dropdown.Option("Cancelado")],
            value="Todos", on_select=self.on_filtro_change, border=ft.InputBorder.NONE, width=140
        )
        
        modalidades = sorted(list(set(t.modalidade for t in self.controller.transmissoes if t.modalidade)))
        self.dd_modalidade = ft.Dropdown(label="Modalidade", options=[ft.dropdown.Option("Todos")] + [ft.dropdown.Option(m) for m in modalidades], value="Todos", on_select=self.on_filtro_change, border=ft.InputBorder.NONE, width=150)
        
        locais = sorted(list(set(t.local for t in self.controller.transmissoes if t.local)))
        self.dd_local = ft.Dropdown(label="Local", options=[ft.dropdown.Option("Todos")] + [ft.dropdown.Option(l) for l in locais], value="Todos", on_select=self.on_filtro_change, border=ft.InputBorder.NONE, width=150)
        
        operadores = sorted(list(set(t.operador for t in self.controller.transmissoes if hasattr(t, 'operador') and t.operador)))
        self.dd_operador = ft.Dropdown(label="Operador", options=[ft.dropdown.Option("Todos")] + [ft.dropdown.Option(o) for o in operadores], value="Todos", on_select=self.on_filtro_change, border=ft.InputBorder.NONE, width=150)

        self.total_row = ft.Container(content=ft.Row([], spacing=30), bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.GREEN_400), border_radius=10, padding=15)
        
        self.btn_exportar = ft.ElevatedButton("Exportar Excel", icon=ft.Icons.FILE_DOWNLOAD, style=ft.ButtonStyle(bgcolor=ft.Colors.GREEN_700, color=ft.Colors.WHITE), on_click=self.exportar_excel, height=45)

        self.controls = [
            ft.Row([ft.IconButton(ft.Icons.ARROW_BACK, on_click=lambda _: self.on_back() if self.on_back else None), ft.Text("📊 Relatórios e Estatísticas", size=30, weight=ft.FontWeight.BOLD), ft.Container(expand=True), self.btn_exportar], alignment=ft.MainAxisAlignment.START, spacing=10),
            ft.Divider(height=10, color=ft.Colors.WHITE10),
            
            # Primeira Linha de Filtros (Busca e Principais)
            ft.Row([self.txt_busca, self.wrap_filter(self.dd_periodo), self.wrap_filter(self.dd_ano)], spacing=10),
            
            # Segunda Linha de Filtros (Categorias)
            ft.Row([
                self.wrap_filter(self.dd_status),
                self.wrap_filter(self.dd_tipo),
                self.wrap_filter(self.dd_modalidade),
                self.wrap_filter(self.dd_local),
                self.wrap_filter(self.dd_operador),
            ], spacing=10, wrap=True),
            
            ft.Divider(height=20, color=ft.Colors.TRANSPARENT),
            ft.Container(content=ft.Row([self.tabela_resumo], scroll=ft.ScrollMode.AUTO), border_radius=15, bgcolor=ft.Colors.with_opacity(0.03, ft.Colors.WHITE), padding=10),
            
            self.total_row,
            
            ft.Text("Lista Detalhada de Transmissões", size=18, weight=ft.FontWeight.BOLD, color=ft.Colors.ORANGE_200),
            ft.Container(content=ft.Row([self.tabela_detalhada], scroll=ft.ScrollMode.AUTO), border_radius=15, bgcolor=ft.Colors.with_opacity(0.03, ft.Colors.WHITE), padding=10),
        ]
        self.popular_tabela()

    def wrap_filter(self, control):
        return ft.Container(content=control, border_radius=12, border=ft.border.all(1, ft.Colors.with_opacity(0.2, ft.Colors.WHITE)), padding=ft.padding.only(left=10, right=10))

    def get_anos_disponiveis(self):
        anos = set(); [anos.add(parse_date(t.data).year) for t in self.controller.transmissoes if parse_date(t.data)]
        return sorted(anos)

    def get_tipos_disponiveis(self):
        tipos = set(); [tipos.add(t.tipo_transmissao) for t in self.controller.transmissoes if t.tipo_transmissao]
        return sorted(tipos)

    def on_filtro_change(self, e):
        self.filtro_periodo = self.dd_periodo.value
        self.filtro_ano = None if self.dd_ano.value == "Todos" else int(self.dd_ano.value)
        self.filtro_tipo = self.dd_tipo.value
        self.filtro_status = self.dd_status.value
        self.filtro_modalidade = self.dd_modalidade.value
        self.filtro_local = self.dd_local.value
        self.filtro_operador = self.dd_operador.value
        self.termo_busca = self.txt_busca.value.lower() if self.txt_busca.value else ""
        self.popular_tabela()
        self.update()

    def get_transmissoes_filtradas(self):
        filtradas = []
        for t in self.controller.transmissoes:
            dt = parse_date(t.data)
            if not dt: continue
            
            # Filtros Fixos
            if self.filtro_ano and dt.year != self.filtro_ano: continue
            if self.filtro_tipo != "Todos" and t.tipo_transmissao != self.filtro_tipo: continue
            if not check_status_match(t.status, self.filtro_status): continue
            if self.filtro_modalidade != "Todos" and t.modalidade != self.filtro_modalidade: continue
            if self.filtro_local != "Todos" and t.local != self.filtro_local: continue
            if self.filtro_operador != "Todos" and getattr(t, 'operador', '') != self.filtro_operador: continue
            
            # Filtro de Busca (Termo)
            if self.termo_busca:
                campos = [t.evento, t.responsavel, format_date_br(t.data), t.local, t.status, t.operador, t.tipo_transmissao, t.modalidade]
                if not any(self.termo_busca in str(c).lower() for c in campos if c): continue
                
            filtradas.append(t)
        return filtradas

    def agrupar_dados(self):
        transmissoes = self.get_transmissoes_filtradas(); grupos = {}
        for t in transmissoes:
            dt = parse_date(t.data)
            if not dt: continue
            if self.filtro_periodo == "Semestral": key = f"{'1º' if dt.month <= 6 else '2º'}/{dt.year}"; sk = (dt.year, 0 if dt.month <= 6 else 1)
            elif self.filtro_periodo == "Anual": key = str(dt.year); sk = (dt.year, 0)
            else: meses = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]; key = f"{meses[dt.month - 1]}/{dt.year}"; sk = (dt.year, dt.month)
            if key not in grupos: grupos[key] = {"sort_key": sk, "transmissoes": 0, "publico": 0, "segundos": 0, "tipos": {}}
            duracao = obter_duracao_efetiva(t)
            grupos[key]["transmissoes"] += 1; grupos[key]["publico"] += t.publico; grupos[key]["segundos"] += converter_tempo_para_segundos(duracao)
            tipo = t.tipo_transmissao or "Outro"; grupos[key]["tipos"][tipo] = grupos[key]["tipos"].get(tipo, 0) + 1
        return sorted(grupos.items(), key=lambda x: x[1]["sort_key"])

    def popular_tabela(self):
        filtradas = self.get_transmissoes_filtradas()
        dados_agrupados = self.agrupar_dados()
        
        # 1. Popular Tabela Resumo
        self.tabela_resumo.columns = [ft.DataColumn(ft.Text("Período")), ft.DataColumn(ft.Text("Público"), numeric=True), ft.DataColumn(ft.Text("Transmissões"), numeric=True), ft.DataColumn(ft.Text("Tipos")), ft.DataColumn(ft.Text("Horas"))]
        self.tabela_resumo.rows = []
        tp, tt, ts, tts = 0, 0, 0, {}
        
        for periodo, info in dados_agrupados:
            tipos_str = ", ".join([f"{t}: {c}" for t, c in sorted(info["tipos"].items())])
            self.tabela_resumo.rows.append(ft.DataRow(cells=[ft.DataCell(ft.Text(periodo)), ft.DataCell(ft.Text(str(info["publico"]))), ft.DataCell(ft.Text(str(info["transmissoes"]))), ft.DataCell(ft.Text(tipos_str, size=12)), ft.DataCell(ft.Text(formatar_segundos_para_tempo(info["segundos"])))]))

        # 2. Popular Tabela Detalhada
        self.tabela_detalhada.rows = []
        for t in filtradas:
            tp += t.publico
            tt += 1
            duracao = obter_duracao_efetiva(t)
            seg = converter_tempo_para_segundos(duracao)
            ts += seg
            tipo = t.tipo_transmissao or "Outro"
            tts[tipo] = tts.get(tipo, 0) + 1
            
            self.tabela_detalhada.rows.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text(format_date_br(t.data), size=11)),
                ft.DataCell(ft.Text(f"{t.horario_inicio}-{t.horario_fim}", size=11)),
                ft.DataCell(ft.Text(t.evento, size=11, width=200)),
                ft.DataCell(ft.Text(t.responsavel, size=11)),
                ft.DataCell(ft.Text(t.tipo_transmissao or "-", size=11)),
                ft.DataCell(ft.Text(t.modalidade or "-", size=11)),
                ft.DataCell(ft.Text(t.local or "-", size=11)),
                ft.DataCell(ft.Text(t.operador or "-", size=11)),
                ft.DataCell(ft.Text(str(t.publico), size=11)),
                ft.DataCell(ft.Text(duracao, size=11)),
                ft.DataCell(ft.Text(t.status, size=10)),
            ]))

        # Adicionar Linha de Somatória na Tabela Detalhada
        if filtradas:
            self.tabela_detalhada.rows.append(
                ft.DataRow(
                    color=ft.Colors.with_opacity(0.1, ft.Colors.CYAN_400),
                    cells=[
                        ft.DataCell(ft.Text("---")),
                        ft.DataCell(ft.Text("---")),
                        ft.DataCell(ft.Text("TOTAL DOS ITENS", weight="bold", color=ft.Colors.CYAN_400)),
                        ft.DataCell(ft.Text("---")),
                        ft.DataCell(ft.Text("---")),
                        ft.DataCell(ft.Text("---")),
                        ft.DataCell(ft.Text("---")),
                        ft.DataCell(ft.Text("---")),
                        ft.DataCell(ft.Text(f"{tp:,}".replace(",", "."), weight="bold", color=ft.Colors.CYAN_400)),
                        ft.DataCell(ft.Text(formatar_segundos_para_tempo(ts), weight="bold", color=ft.Colors.CYAN_400)),
                        ft.DataCell(ft.Text("---")),
                    ]
                )
            )

        # 3. Atualizar Barra de Totais
        total_tipos_str = ", ".join([f"{t}: {c}" for t, c in sorted(tts.items())])
        self.total_row.content = ft.Row([
            ft.Container(content=ft.Row([ft.Icon(ft.Icons.SUMMARIZE, color=ft.Colors.GREEN_400), ft.Text("TOTAL GERAL", weight="bold")]), width=150),
            ft.Container(content=ft.Column([ft.Text("Público Total", size=12), ft.Text(f"{tp:,}".replace(",", "."), size=18, weight="bold")], spacing=2), width=130),
            ft.Container(content=ft.Column([ft.Text("Transmissões", size=12), ft.Text(str(tt), size=18, weight="bold")], spacing=2), width=130),
            ft.Container(content=ft.Column([ft.Text("Tipos", size=12), ft.Text(total_tipos_str, size=11)], spacing=2), expand=True),
            ft.Container(content=ft.Column([ft.Text("Horas Totais", size=12), ft.Text(formatar_segundos_para_tempo(ts), size=18, weight="bold", color=ft.Colors.CYAN_400)], spacing=2), width=150)
        ], spacing=20)

    def abrir_pasta(self, caminho):
        try:
            pasta = os.path.dirname(caminho)
            if os.name == 'nt': os.startfile(pasta)
            else: import subprocess; subprocess.run(['open' if os.uname().sysname == 'Darwin' else 'xdg-open', pasta])
        except: pass

    def exportar_excel(self, e):
        def _get_path():
            root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True); root.focus_force()
            path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile="relatorio.xlsx")
            root.destroy(); return path
        caminho = _get_path()
        if not caminho: return
        
        self.btn_exportar.disabled = True; self.btn_exportar.text = "Exportando..."; self.update()
        
        async def _exec():
            try:
                def _build():
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
                    wb = Workbook()
                    ws_res = wb.active
                    ws_res.title = "Resumo"
                    
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    # --- ABA RESUMO ---
                    headers_res = ["Período", "Público", "Transmissões", "Tipos", "Horas"]
                    for col, h in enumerate(headers_res, 1):
                        cell = ws_res.cell(1, col, h)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", start_color="1F4E79")
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = thin_border
                    
                    dados_agrupados = self.agrupar_dados()
                    for row, (periodo, info) in enumerate(dados_agrupados, 2):
                        ws_res.cell(row, 1, periodo).border = thin_border
                        ws_res.cell(row, 2, info["publico"]).border = thin_border
                        ws_res.cell(row, 3, info["transmissoes"]).border = thin_border
                        ws_res.cell(row, 4, ", ".join([f"{t}: {c}" for t, c in info["tipos"].items()])).border = thin_border
                        ws_res.cell(row, 5, formatar_segundos_para_tempo(info["segundos"])).border = thin_border

                    # Ajuste de largura Resumo
                    for col in ws_res.columns:
                        ws_res.column_dimensions[col[0].column_letter].width = 20

                    # --- ABA DETALHADA ---
                    ws_det = wb.create_sheet("Lista Detalhada")
                    headers_det = ["Data", "Horário", "Evento", "Responsável", "Tipo", "Modalidade", "Local", "Operador", "Público", "Duração", "Status"]
                    for col, h in enumerate(headers_det, 1):
                        cell = ws_det.cell(1, col, h)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", start_color="C65911")
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = thin_border
                    
                    filtradas = self.get_transmissoes_filtradas()
                    for row, t in enumerate(filtradas, 2):
                        duracao = obter_duracao_efetiva(t)
                        data = [format_date_br(t.data), f"{t.horario_inicio} às {t.horario_fim}", t.evento, t.responsavel, 
                                t.tipo_transmissao or "-", t.modalidade or "-", t.local or "-", t.operador or "-", 
                                t.publico, duracao, t.status]
                        for col, val in enumerate(data, 1):
                            cell = ws_det.cell(row, col, val)
                            cell.border = thin_border
                            if col in [1, 2, 9, 10]: cell.alignment = Alignment(horizontal="center")

                    # Linha de Total no Excel Detalhado
                    last_row = len(filtradas) + 2
                    total_publico = sum(t.publico for t in filtradas)
                    total_seg = sum(converter_tempo_para_segundos(obter_duracao_efetiva(t)) for t in filtradas)
                    
                    for col in range(1, 12):
                        cell = ws_det.cell(last_row, col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill("solid", start_color="FCE4D6")
                        cell.border = thin_border
                    
                    ws_det.cell(last_row, 3, "TOTAL GERAL").alignment = Alignment(horizontal="right")
                    ws_det.cell(last_row, 9, total_publico).alignment = Alignment(horizontal="center")
                    ws_det.cell(last_row, 10, formatar_segundos_para_tempo(total_seg)).alignment = Alignment(horizontal="center")

                    # Ajuste Automático de Colunas Detalhado
                    for col in ws_det.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                            except: pass
                        ws_det.column_dimensions[column].width = min(max_length + 2, 60)

                    wb.save(caminho)
                
                await asyncio.to_thread(_build)
                self.show_popup("Relatório completo exportado!", caminho)
            except Exception as ex:
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Erro ao exportar Excel: {ex}"), bgcolor=ft.Colors.RED_700); self.page.snack_bar.open = True
            finally:
                self.btn_exportar.disabled = False; self.btn_exportar.text = "Exportar Excel"; self.page.update()
        self.page.run_task(_exec)

    def show_popup(self, text, path):
        def on_open_click(e):
            self.abrir_pasta(path); dlg.open = False; self.page.update()
        dlg = ft.AlertDialog(
            title=ft.Text("Arquivo Salvo"),
            content=ft.Text(f"{text}\nDeseja abrir a pasta agora?"),
            actions=[
                ft.TextButton("Não", on_click=lambda _: self.fechar_dlg(dlg)),
                ft.ElevatedButton("Abrir Pasta", icon=ft.Icons.FOLDER_OPEN, bgcolor=ft.Colors.GREEN_700, on_click=on_open_click)
            ]
        )
        self.page.overlay.append(dlg); dlg.open = True; self.page.update()

    def fechar_dlg(self, dlg): dlg.open = False; self.page.update()
