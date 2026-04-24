import flet as ft
import os
from datetime import datetime
from utils.helpers import normalize_date, format_date_br, check_status_match, get_status_info, parse_date
from utils.dialog_helper import mostrar_detalhes_transmissao, abrir_menu_contexto

class HistoricoView(ft.Column):
    def __init__(self, controller, on_edit):
        super().__init__(expand=True, spacing=20)
        self.controller = controller
        self.on_edit = on_edit
        self.init_ui()

    def init_ui(self):
        from utils.helpers import get_status_options
        self.sort_column_index = 0
        self.sort_ascending = False

        self.txt_busca = ft.TextField(
            label="Pesquisar no Histórico", 
            prefix_icon=ft.Icons.SEARCH, 
            on_change=lambda _: self.filtrar(), 
            expand=True, height=45, text_size=13,
            bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.WHITE),
            border_radius=10,
            border_color=ft.Colors.WHITE10,
            focused_border_color=ft.Colors.CYAN_400
        )
        self.dd_status = ft.Dropdown(label="Status", options=[ft.dropdown.Option("Todos")] + [ft.dropdown.Option(o) for o in get_status_options()], value="Todos", on_select=lambda _: self.filtrar(), width=180, height=45, text_size=12)
        tipos = sorted(list(set(t.tipo_transmissao for t in self.controller.transmissoes if t.tipo_transmissao)))
        self.dd_tipo = ft.Dropdown(label="Tipo", options=[ft.dropdown.Option("Todos")] + [ft.dropdown.Option(o) for o in tipos], value="Todos", on_select=lambda _: self.filtrar(), width=150, height=45, text_size=12)
        
        modalidades = sorted(list(set(t.modalidade for t in self.controller.transmissoes if t.modalidade)))
        self.dd_modalidade = ft.Dropdown(label="Modalidade", options=[ft.dropdown.Option("Todos")] + [ft.dropdown.Option(o) for o in modalidades], value="Todos", on_select=lambda _: self.filtrar(), width=150, height=45, text_size=12)
        
        locais = sorted(list(set(t.local for t in self.controller.transmissoes if t.local)))
        self.dd_local = ft.Dropdown(label="Local", options=[ft.dropdown.Option("Todos")] + [ft.dropdown.Option(l) for l in locais], value="Todos", on_select=lambda _: self.filtrar(), width=180, height=45, text_size=12)
        
        operadores = sorted(list(set(t.operador for t in self.controller.transmissoes if hasattr(t, 'operador') and t.operador)))
        self.dd_operador = ft.Dropdown(label="Operador", options=[ft.dropdown.Option("Todos")] + [ft.dropdown.Option(o) for o in operadores], value="Todos", on_select=lambda _: self.filtrar(), width=180, height=45, text_size=12)

        self.dd_periodo_tipo = ft.Dropdown(
            label="Tipo de Período",
            options=[
                ft.dropdown.Option("Todos"),
                ft.dropdown.Option("Por Mês"),
                ft.dropdown.Option("Por Ano"),
                ft.dropdown.Option("Personalizado")
            ],
            value="Por Mês",
            on_select=self.on_periodo_change,
            width=180, height=45, text_size=12
        )
        
        meses_nomes = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        self.dd_filtro_mes = ft.Dropdown(
            label="Mês",
            options=[ft.dropdown.Option(text=m, key=str(i+1)) for i, m in enumerate(meses_nomes)],
            value=str(datetime.now().month),
            on_select=lambda _: self.filtrar(),
            width=140, height=45, text_size=12, visible=False
        )
        
        anos_disp = sorted(list(set(parse_date(t.data).year for t in self.controller.transmissoes if parse_date(t.data))), reverse=True)
        if not anos_disp: anos_disp = [datetime.now().year]
        
        self.dd_filtro_ano = ft.Dropdown(
            label="Ano",
            options=[ft.dropdown.Option(str(a)) for a in anos_disp],
            value=str(datetime.now().year),
            on_select=lambda _: self.filtrar(),
            width=110, height=45, text_size=12, visible=False
        )
        
        self.txt_data_inicio = ft.TextField(
            label="De:", hint_text="dd/mm/aaaa", 
            width=130, height=45, text_size=12, 
            on_submit=lambda _: self.filtrar(), 
            visible=False,
            bgcolor=ft.Colors.with_opacity(0.15, ft.Colors.WHITE),
            border_radius=8,
            border_color=ft.Colors.WHITE10
        )
        self.txt_data_fim = ft.TextField(
            label="Até:", hint_text="dd/mm/aaaa", 
            width=130, height=45, text_size=12, 
            on_submit=lambda _: self.filtrar(), 
            visible=False,
            bgcolor=ft.Colors.with_opacity(0.15, ft.Colors.WHITE),
            border_radius=8,
            border_color=ft.Colors.WHITE10
        )

        self.row_filtros = ft.Row([
            self.dd_periodo_tipo,
            self.dd_filtro_mes,
            self.dd_filtro_ano,
            self.txt_data_inicio,
            self.txt_data_fim,
            self.dd_status, 
            self.dd_tipo, 
            self.dd_modalidade, 
            self.dd_local, 
            self.dd_operador
        ], spacing=10, visible=True, wrap=True)
        
        self.tabela = ft.DataTable(
            sort_column_index=self.sort_column_index,
            sort_ascending=self.sort_ascending,
            columns=[
                ft.DataColumn(ft.Text(""), visible=False), # Checkbox Column (index 0)
                ft.DataColumn(ft.Text("Data"), on_sort=self.on_sort_click),
                ft.DataColumn(ft.Text("Horário"), on_sort=self.on_sort_click),
                ft.DataColumn(ft.Text("Evento"), on_sort=self.on_sort_click),
                ft.DataColumn(ft.Text("Responsável"), on_sort=self.on_sort_click),
                ft.DataColumn(ft.Text("Tipo"), on_sort=self.on_sort_click),
                ft.DataColumn(ft.Text("Modalidade"), on_sort=self.on_sort_click),
                ft.DataColumn(ft.Text("Local"), on_sort=self.on_sort_click),
                ft.DataColumn(ft.Text("Operador"), on_sort=self.on_sort_click),
                ft.DataColumn(ft.Text("Status"), on_sort=self.on_sort_click),
                ft.DataColumn(ft.Text("Ações")),
            ],
            column_spacing=20,
            rows=[],
            heading_row_color=ft.Colors.with_opacity(0.1, ft.Colors.BLUE_400),
            border_radius=10,
        )

        # Configuração da Barra de Meses (Visual)
        self.selection_mode = False
        self.selected_ids = set()
        self.meses_nomes = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        self.row_meses_disponiveis = ft.Row(spacing=15)
        
        self.txt_mes_selecionado = ft.Text(self.meses_nomes[datetime.now().month-1], color=ft.Colors.WHITE, size=18, weight=ft.FontWeight.BOLD)
        self.menu_mes = ft.PopupMenuButton(
            content=ft.Row([self.txt_mes_selecionado, ft.Icon(ft.Icons.ARROW_DROP_DOWN, color=ft.Colors.WHITE, size=20)], spacing=5),
            items=[ft.PopupMenuItem(content=ft.Text("Todos"), on_click=lambda e: self.set_filtro_mes_bar(0, "Todos"))] + \
                  [ft.PopupMenuItem(content=ft.Text(m), on_click=lambda e, m=m, i=i: self.set_filtro_mes_bar(i+1, m)) for i, m in enumerate(self.meses_nomes)]
        )
        
        self.txt_ano_selecionado = ft.Text(str(datetime.now().year), color=ft.Colors.WHITE, size=18, weight=ft.FontWeight.BOLD)
        self.menu_ano = ft.PopupMenuButton(
            content=ft.Row([self.txt_ano_selecionado, ft.Icon(ft.Icons.ARROW_DROP_DOWN, color=ft.Colors.WHITE, size=20)], spacing=5),
            items=[]
        )

        self.filter_bar = ft.Container(
            content=ft.Row([
                ft.Container(content=ft.Row([ft.Text("MESES:", color=ft.Colors.WHITE60, size=11, weight=ft.FontWeight.BOLD), self.row_meses_disponiveis], spacing=15), expand=True),
                ft.Container(content=self.menu_mes, alignment=ft.Alignment(0, 0)),
                ft.Container(
                    content=self.menu_ano, 
                    alignment=ft.Alignment(1, 0),
                    border=ft.border.all(1, ft.Colors.WHITE38),
                    border_radius=8,
                    padding=ft.padding.symmetric(horizontal=12, vertical=4),
                    bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.BLACK)
                )
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            gradient=ft.LinearGradient(
                begin=ft.Alignment(-1, 0),
                end=ft.Alignment(1, 0),
                colors=[ft.Colors.GREEN_900, ft.Colors.GREEN_600]
            ),
            padding=ft.padding.symmetric(horizontal=20, vertical=8),
            border_radius=10,
            margin=ft.margin.only(bottom=5),
            visible=True # Fica visível pois o padrão agora é "Por Mês"
        )

        self.btn_exportar_links = ft.ElevatedButton("Exportar Links", icon=ft.Icons.LINK, on_click=self.exportar_links, style=ft.ButtonStyle(color=ft.Colors.WHITE, bgcolor=ft.Colors.BLUE_700))
        self.btn_cancelar_links = ft.TextButton("Cancelar", icon=ft.Icons.CLOSE, on_click=self.cancelar_selecao_links, visible=False, style=ft.ButtonStyle(color=ft.Colors.RED_400))
        self.selection_bar_control = ft.Container(visible=False) # Placeholder para a barra de "Selecionar Todos"

        self.controls = [
            ft.Row([
                ft.Text("📜 Histórico de Transmissões", size=30, weight=ft.FontWeight.BOLD),
                ft.Container(expand=True),
                ft.Row([
                    self.btn_cancelar_links,
                    self.btn_exportar_links,
                    ft.ElevatedButton(
                        "Modificações", 
                        icon=ft.Icons.HISTORY, 
                        on_click=lambda _: self.abrir_historico_modificacoes(),
                        style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_GREY_800, color=ft.Colors.WHITE)
                    ),
                    ft.ElevatedButton(
                        "Exportar", 
                        icon=ft.Icons.DOWNLOAD, 
                        on_click=lambda _: self.abrir_opcoes_exportar(),
                        style=ft.ButtonStyle(bgcolor=ft.Colors.INDIGO_900, color=ft.Colors.WHITE)
                    ),
                ], spacing=10)
            ], alignment=ft.MainAxisAlignment.START),
            ft.Row([
                self.txt_busca, 
                ft.ElevatedButton(
                    "Filtros", 
                    icon=ft.Icons.FILTER_LIST, 
                    on_click=lambda _: self.toggle_filtros(),
                    style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_900, color=ft.Colors.WHITE, shape=ft.RoundedRectangleBorder(radius=8))
                )
            ], spacing=10),
            self.row_filtros,
            self.filter_bar,
            self.selection_bar_control,
            ft.Container(
                content=ft.Column([
                    ft.Row([self.tabela], scroll=ft.ScrollMode.AUTO)
                ], scroll=ft.ScrollMode.ADAPTIVE),
                padding=10, 
                bgcolor=ft.Colors.WHITE10, 
                border_radius=15, 
                expand=True
            )
        ]
        self.update_year_menu()
        self.update_meses_bar()
        self.filtrar()
        self.filtrar()

    def on_sort_click(self, e):
        if e.column_index == 0: return # Ignora clique na coluna de checkbox
        self.sort_column_index = e.column_index
        self.sort_ascending = not self.sort_ascending
        self.tabela.sort_column_index = self.sort_column_index
        self.tabela.sort_ascending = self.sort_ascending
        self.filtrar()

    def toggle_filtros(self):
        self.row_filtros.visible = not self.row_filtros.visible
        self.update()

    def on_periodo_change(self, e):
        tipo = self.dd_periodo_tipo.value
        self.dd_filtro_mes.visible = False # (Ocultamos os dropdowns pois agora usamos a barra visual)
        self.dd_filtro_ano.visible = False
        self.txt_data_inicio.visible = (tipo == "Personalizado")
        self.txt_data_fim.visible = (tipo == "Personalizado")
        self.filter_bar.visible = (tipo in ["Por Mês", "Por Ano"])
        self.filtrar()
        self.update()

    def filtrar(self):
        termo = self.txt_busca.value.lower() if self.txt_busca.value else ""
        status = self.dd_status.value
        tipo = self.dd_tipo.value
        modalidade = self.dd_modalidade.value
        local = self.dd_local.value
        operador = self.dd_operador.value
        periodo_tipo = self.dd_periodo_tipo.value
        mes_f = int(self.dd_filtro_mes.value) if self.dd_filtro_mes.value else None
        ano_f = int(self.dd_filtro_ano.value) if self.dd_filtro_ano.value else None
        
        from utils.helpers import parse_date
        hoje = datetime.now()
        hoje_str = hoje.strftime("%Y-%m-%d")
        
        # Datas para o filtro personalizado
        dt_ini = None; dt_fim = None
        if periodo_tipo == "Personalizado":
            if self.txt_data_inicio.value: dt_ini = parse_date(self.txt_data_inicio.value)
            if self.txt_data_fim.value: dt_fim = parse_date(self.txt_data_fim.value)
            
            # Se o usuário digitou algo mas o parse falhou, avisamos
            if (self.txt_data_inicio.value and not dt_ini) or (self.txt_data_fim.value and not dt_fim):
                self.page.snack_bar = ft.SnackBar(ft.Text("Formato de data inválido! Use DD/MM/AAAA"), bgcolor=ft.Colors.ORANGE_900)
                self.page.snack_bar.open = True
                if self.page: self.page.update()
        self.tabela.rows = []
        eventos = []
        for t in self.controller.transmissoes:
            dt_t = parse_date(t.data)
            if not dt_t: continue
            dt_str = normalize_date(t.data)

            # 1. Filtro de Período (Prioritário)
            if periodo_tipo == "Por Mês":
                if dt_t.month != mes_f or dt_t.year != ano_f: continue
            elif periodo_tipo == "Por Ano":
                if dt_t.year != ano_f: continue
            elif periodo_tipo == "Personalizado":
                if dt_ini and dt_t < dt_ini: continue
                if dt_fim and dt_t > dt_fim: continue
            else:
                # Se for "Todos", respeita a regra padrão de não mostrar futuro no histórico
                # a menos que haja outros filtros ativos
                if not termo and status == "Todos" and tipo == "Todos" and modalidade == "Todos" and local == "Todos" and operador == "Todos":
                    if dt_str > hoje_str: continue

            if termo:
                # Localiza em todos os campos relevantes
                campos = [
                    t.evento,
                    t.responsavel,
                    format_date_br(t.data),
                    f"{t.horario_inicio} às {t.horario_fim}",
                    t.tipo_transmissao,
                    t.modalidade,
                    t.local,
                    t.status,
                    getattr(t, 'operador', '')
                ]
                if not any(termo in str(c).lower() for c in campos if c):
                    continue

            if not check_status_match(t.status, status): continue
            if tipo != "Todos" and t.tipo_transmissao != tipo: continue
            if modalidade != "Todos" and t.modalidade != modalidade: continue
            if local != "Todos" and t.local != local: continue
            if operador != "Todos" and getattr(t, 'operador', '') != operador: continue
            eventos.append(t)

        def get_sort_key(t):
            if self.sort_column_index == 0: return (normalize_date(t.data), t.horario_inicio)
            if self.sort_column_index == 1: return t.horario_inicio
            if self.sort_column_index == 2: return t.evento.lower()
            if self.sort_column_index == 3: return t.responsavel.lower()
            if self.sort_column_index == 4: return t.tipo_transmissao.lower() if t.tipo_transmissao else ""
            if self.sort_column_index == 5: return t.modalidade.lower() if t.modalidade else ""
            if self.sort_column_index == 6: return t.local.lower() if t.local else ""
            if self.sort_column_index == 7: return t.operador.lower() if hasattr(t, 'operador') and t.operador else ""
            if self.sort_column_index == 8: return t.status.lower()
            return (normalize_date(t.data), t.horario_inicio)

        eventos.sort(key=get_sort_key, reverse=not self.sort_ascending)
        
        # Barra de seleção "Todos" no topo da tabela
        if self.selection_mode:
            todos_ids = set(t.id for t in eventos)
            todos_sel = len(todos_ids) > 0 and todos_ids.issubset(self.selected_ids)
            alguns_sel = len(self.selected_ids) > 0 and not todos_sel
            
            cb_todos = ft.Checkbox(
                value=todos_sel,
                tristate=True if alguns_sel else False,
                fill_color=ft.Colors.CYAN_400,
                on_change=lambda e: self.toggle_select_all(e.control.value),
            )
            if alguns_sel: cb_todos.value = None
            
            self.selection_bar_control.content = ft.Container(
                content=ft.Row([
                    cb_todos,
                    ft.Text("Selecionar Todos", size=14, weight=ft.FontWeight.W_500),
                    ft.Container(expand=True),
                    ft.Container(
                        content=ft.Text(f"{len(self.selected_ids)} selecionados", size=12, color=ft.Colors.WHITE70),
                        padding=ft.padding.symmetric(horizontal=12, vertical=4),
                        bgcolor=ft.Colors.with_opacity(0.15, ft.Colors.CYAN),
                        border_radius=10
                    )
                ], spacing=10),
                padding=ft.padding.symmetric(horizontal=15, vertical=5),
                bgcolor=ft.Colors.with_opacity(0.05, ft.Colors.CYAN),
                border_radius=10,
                margin=ft.margin.only(bottom=10),
                visible=True
            )
            self.selection_bar_control.visible = True
        else:
            self.selection_bar_control.visible = False

        for t in eventos:
            status_info = get_status_info(t.status)
            cells = [
                ft.DataCell(
                    ft.Checkbox(
                        value=t.id in self.selected_ids,
                        on_change=lambda e, tid=t.id: self.toggle_selection(tid, e.control.value),
                        fill_color=ft.Colors.CYAN_400
                    )
                ) if self.selection_mode else ft.DataCell(ft.Container(), visible=False),
                ft.DataCell(ft.Text(format_date_br(t.data), size=12)),
                ft.DataCell(ft.Text(f"{t.horario_inicio} às {t.horario_fim}", size=12, weight=ft.FontWeight.BOLD, no_wrap=True)),
                ft.DataCell(
                    ft.Container(
                        content=ft.GestureDetector(
                            content=ft.Text(t.evento, size=13, weight=ft.FontWeight.W_500, color=ft.Colors.BLUE_200),
                            on_tap=lambda _, item=t: mostrar_detalhes_transmissao(self.page, item, self.controller, self.atualizar),
                            on_secondary_tap=lambda e, item=t: abrir_menu_contexto(self.page, item, self.controller, self.on_edit, self.confirmar_exclusao, self.atualizar),
                        ),
                        width=300, 
                        padding=ft.padding.symmetric(vertical=5)
                    )
                ),
                ft.DataCell(ft.Text(t.responsavel, size=12)),
                ft.DataCell(ft.Text(t.tipo_transmissao or "-", size=12)),
                ft.DataCell(ft.Text(t.modalidade or "-", size=12)),
                ft.DataCell(ft.Text(t.local or "-", size=11, color=ft.Colors.WHITE38)),
                ft.DataCell(ft.Text(t.operador or "-", size=12)),
                ft.DataCell(ft.Container(content=ft.Text(status_info["label"], size=10, weight=ft.FontWeight.BOLD, color=ft.Colors.BLACK), bgcolor=status_info["color"], padding=ft.padding.symmetric(horizontal=10, vertical=4), border_radius=8)),
                ft.DataCell(ft.Row([
                    ft.IconButton(ft.Icons.INFO_OUTLINE, icon_size=18, icon_color=ft.Colors.CYAN_200, on_click=lambda _, item=t: mostrar_detalhes_transmissao(self.page, item, self.controller, self.atualizar)),
                    ft.IconButton(ft.Icons.EDIT, icon_size=18, icon_color=ft.Colors.BLUE_200, on_click=lambda _, item=t: self.on_edit(item)),
                    ft.IconButton(ft.Icons.DELETE, icon_size=18, icon_color=ft.Colors.RED_200, on_click=lambda _, item=t: self.confirmar_exclusao(item)),
                ], spacing=0)),
            ]
            self.tabela.rows.append(ft.DataRow(cells=cells))
        try:
            self.update()
        except: pass

    def atualizar(self): 
        # Atualiza as opções dos dropdowns antes de filtrar
        tipos = sorted(list(set(t.tipo_transmissao for t in self.controller.transmissoes if t.tipo_transmissao)))
        self.dd_tipo.options = [ft.dropdown.Option("Todos")] + [ft.dropdown.Option(o) for o in tipos]
        
        modalidades = sorted(list(set(t.modalidade for t in self.controller.transmissoes if t.modalidade)))
        self.dd_modalidade.options = [ft.dropdown.Option("Todos")] + [ft.dropdown.Option(o) for o in modalidades]
        
        locais = sorted(list(set(t.local for t in self.controller.transmissoes if t.local)))
        self.dd_local.options = [ft.dropdown.Option("Todos")] + [ft.dropdown.Option(l) for l in locais]
        
        operadores = sorted(list(set(t.operador for t in self.controller.transmissoes if hasattr(t, 'operador') and t.operador)))
        self.dd_operador.options = [ft.dropdown.Option("Todos")] + [ft.dropdown.Option(o) for o in operadores]
        
        # Manter o valor atual se ele ainda existir, senão volta para "Todos"
        for dd in [self.dd_tipo, self.dd_modalidade, self.dd_local, self.dd_operador]:
            valid_keys = [str(o.key) for o in dd.options if o.key is not None]
            valid_texts = [str(o.text) for o in dd.options if o.text is not None]
            if dd.value not in valid_keys and dd.value not in valid_texts:
                dd.value = "Todos"
        
        self.filtrar()

    def confirmar_exclusao(self, t):
        dlg = ft.AlertDialog(title=ft.Text("Confirmar Exclusão"), content=ft.Text(f"Deseja excluir '{t.evento}'?"), actions=[ft.TextButton("Cancelar", on_click=lambda _: self.fechar_dlg(dlg)), ft.ElevatedButton("Excluir", bgcolor=ft.Colors.RED_700, on_click=lambda _: self.deletar_e_fechar(t, dlg))])
        self.page.overlay.append(dlg); dlg.open = True; self.page.update()
    
    def fechar_dlg(self, dlg): dlg.open = False; self.page.update()
    def deletar_e_fechar(self, t, dlg): self.controller.deletar(t.id); dlg.open = False; self.filtrar()

    def abrir_historico_modificacoes(self):
        historico_original = self.controller.get_historico()
        self.hist_data = historico_original[:] # Cópia para manipulação
        self.hist_sort_column = 0
        self.hist_sort_ascending = False
        
        def filtrar_e_atualizar():
            try:
                busca = txt_busca_hist.value.lower() if txt_busca_hist.value else ""
                user_f = dd_usuario.value
                acao_f = dd_acao.value
                
                dados = []
                for h in historico_original:
                    if busca:
                        full_text = f"{h.get('data_hora','')} {h.get('usuario','')} {h.get('acao','')} {h.get('detalhes','')}".lower()
                        if busca not in full_text: continue
                    
                    if user_f != "Todos" and h.get("usuario") != user_f: continue
                    if acao_f != "Todos" and h.get("acao") != acao_f: continue
                    
                    dados.append(h)

                def get_val(item):
                    cols = ["data_hora", "usuario", "acao", "detalhes"]
                    key = cols[self.hist_sort_column]
                    val = item.get(key, "")
                    if key == "data_hora":
                        try: return datetime.strptime(val, "%d/%m/%Y %H:%M:%S")
                        except: return val
                    return str(val).lower()

                dados.sort(key=get_val, reverse=not self.hist_sort_ascending)
                self.hist_data = dados

                tabela_hist.rows = [
                    ft.DataRow(
                        cells=[
                            ft.DataCell(ft.Text(h.get("data_hora", ""), size=12, color=ft.Colors.WHITE70)),
                            ft.DataCell(ft.Text(h.get("usuario", ""), size=12, weight=ft.FontWeight.W_500)),
                            ft.DataCell(
                                ft.Container(
                                    content=ft.Text(
                                        h.get("acao", ""), 
                                        size=11, 
                                        weight=ft.FontWeight.BOLD, 
                                        color=ft.Colors.BLACK
                                    ),
                                    bgcolor=ft.Colors.BLUE_400 if h.get("acao") == "ALTERAÇÃO" else ft.Colors.GREEN_400 if h.get("acao") == "INSERÇÃO" else ft.Colors.RED_400,
                                    padding=ft.padding.symmetric(horizontal=8, vertical=2),
                                    border_radius=5
                                )
                            ),
                            ft.DataCell(
                                ft.Container(
                                    content=ft.Text(h.get("detalhes", ""), size=11, no_wrap=True, overflow=ft.TextOverflow.ELLIPSIS), 
                                    width=550,
                                    tooltip="Clique para ver detalhes completos",
                                    on_click=lambda _, val=h.get("detalhes", ""): self.mostrar_full_detalhes(val)
                                )
                            ),
                        ]
                    ) for h in dados
                ]
                try:
                    if tabela_hist.page: tabela_hist.update()
                    if dlg.page: dlg.update()
                except: pass
            except Exception as ex:
                print(f"Erro ao filtrar histórico: {ex}")

        def sort_hist(column_index):
            if self.hist_sort_column == column_index:
                self.hist_sort_ascending = not self.hist_sort_ascending
            else:
                self.hist_sort_column = column_index
                self.hist_sort_ascending = True
            
            tabela_hist.sort_column_index = column_index
            tabela_hist.sort_ascending = self.hist_sort_ascending
            filtrar_e_atualizar()

        def limpar_filtros_hist(e):
            txt_busca_hist.value = ""
            dd_usuario.value = "Todos"
            dd_acao.value = "Todos"
            filtrar_e_atualizar()

        # Controles de Filtro
        txt_busca_hist = ft.TextField(
            label="Pesquisar no histórico...",
            prefix_icon=ft.Icons.SEARCH,
            on_change=lambda e: filtrar_e_atualizar(),
            expand=True,
            height=45,
            text_size=13,
            border_radius=10,
            bgcolor=ft.Colors.with_opacity(0.05, ft.Colors.WHITE)
        )
        
        usuarios = sorted(list(set(h.get("usuario", "Desconhecido") for h in historico_original)))
        dd_usuario = ft.Dropdown(
            label="Usuário",
            options=[ft.dropdown.Option("Todos")] + [ft.dropdown.Option(u) for u in usuarios],
            value="Todos",
            on_select=lambda _: filtrar_e_atualizar(),
            width=150, height=45, text_size=12
        )

        acoes = sorted(list(set(h.get("acao", "") for h in historico_original)))
        dd_acao = ft.Dropdown(
            label="Ação",
            options=[ft.dropdown.Option("Todos")] + [ft.dropdown.Option(a) for a in acoes],
            value="Todos",
            on_select=lambda _: filtrar_e_atualizar(),
            width=150, height=45, text_size=12
        )

        tabela_hist = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Data/Hora"), on_sort=lambda e: sort_hist(e.column_index)),
                ft.DataColumn(ft.Text("Usuário"), on_sort=lambda e: sort_hist(e.column_index)),
                ft.DataColumn(ft.Text("Ação"), on_sort=lambda e: sort_hist(e.column_index)),
                ft.DataColumn(ft.Text("Detalhes"), on_sort=lambda e: sort_hist(e.column_index)),
            ],
            rows=[],
            column_spacing=20,
            heading_row_color=ft.Colors.with_opacity(0.1, ft.Colors.BLUE_GREY_400),
            heading_row_height=45,
            data_row_min_height=40,
            data_row_max_height=60,
        )

        dlg = ft.AlertDialog(
            title=ft.Row([
                ft.Icon(ft.Icons.HISTORY, color=ft.Colors.BLUE_400), 
                ft.Text("Histórico de Modificações", size=22, weight=ft.FontWeight.BOLD)
            ], alignment=ft.MainAxisAlignment.START),
            content=ft.Container(
                content=ft.Column([
                    ft.Row([
                        txt_busca_hist, 
                        dd_usuario, 
                        dd_acao,
                        ft.IconButton(ft.Icons.FILTER_ALT_OFF, tooltip="Limpar Filtros", on_click=limpar_filtros_hist, icon_color=ft.Colors.WHITE38),
                        ft.VerticalDivider(width=1, color=ft.Colors.WHITE10),
                        ft.IconButton(ft.Icons.TABLE_CHART, tooltip="Exportar Excel", on_click=lambda _: self.exportar_historico_excel(self.hist_data), icon_color=ft.Colors.GREEN_400),
                        ft.IconButton(ft.Icons.TEXT_SNIPPET, tooltip="Exportar TXT", on_click=lambda _: self.exportar_historico_txt(self.hist_data), icon_color=ft.Colors.BLUE_200),
                    ], spacing=10),
                    ft.Divider(height=1, color=ft.Colors.WHITE10),
                    # AQUI: Column com scroll=AUTO garante a rolagem vertical da tabela
                    ft.Column([
                        ft.Row([tabela_hist], scroll=ft.ScrollMode.AUTO)
                    ], scroll=ft.ScrollMode.AUTO, expand=True, spacing=0)
                ], spacing=15),
                width=1200, 
                height=self.page.height * 0.8 if self.page else 750,
                padding=ft.padding.only(bottom=10)
            ),
            actions=[
                ft.TextButton("Fechar", on_click=lambda _: self.fechar_dlg(dlg), style=ft.ButtonStyle(color=ft.Colors.BLUE_200))
            ],
            actions_alignment=ft.MainAxisAlignment.END,
            shape=ft.RoundedRectangleBorder(radius=15),
        )
        
        self.page.overlay.append(dlg)
        dlg.open = True
        filtrar_e_atualizar()
        self.page.update()

    def mostrar_full_detalhes(self, texto):
        dlg = ft.AlertDialog(
            title=ft.Row([ft.Icon(ft.Icons.INFO_OUTLINE, color=ft.Colors.BLUE_400), ft.Text("Detalhes da Modificação")]),
            content=ft.Container(
                content=ft.Column([
                    ft.Text(texto, size=14, selectable=True),
                ], scroll=ft.ScrollMode.AUTO, tight=True),
                width=700,
                height=500,
                padding=10
            ),
            actions=[ft.TextButton("Fechar", on_click=lambda _: self.fechar_dlg(dlg))],
            shape=ft.RoundedRectangleBorder(radius=12)
        )
        self.page.overlay.append(dlg); dlg.open = True; self.page.update()

    def exportar_historico_excel(self, dados):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from tkinter import filedialog
        import tkinter as tk

        if not dados:
            self.page.snack_bar = ft.SnackBar(ft.Text("Nenhum dado para exportar!"), bgcolor=ft.Colors.ORANGE_800)
            self.page.snack_bar.open = True; self.page.update()
            return

        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile=f"log_modificacoes_{datetime.now().strftime('%Y%m%d')}.xlsx")
        root.destroy()

        if path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Histórico de Modificações"
                
                # Cabeçalho
                headers = ["Data/Hora", "Usuário", "Ação", "Detalhes"]
                ws.append(headers)
                
                # Estilo Cabeçalho
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # Dados
                for h in dados:
                    ws.append([h.get("data_hora"), h.get("usuario"), h.get("acao"), h.get("detalhes")])

                # Ajuste de largura das colunas
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try: max_length = max(max_length, len(str(cell.value)))
                        except: pass
                    ws.column_dimensions[column].width = min(max_length + 2, 100)

                wb.save(path)
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Log exportado com sucesso: {os.path.basename(path)}"), bgcolor=ft.Colors.GREEN_700)
            except Exception as e:
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Erro ao exportar Excel: {e}"), bgcolor=ft.Colors.RED_700)
            self.page.snack_bar.open = True; self.page.update()

    def exportar_historico_txt(self, dados):
        from tkinter import filedialog
        import tkinter as tk
        if not dados: return
        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Arquivo de Texto", "*.txt")], initialfile=f"log_modificacoes_{datetime.now().strftime('%Y%m%d')}.txt")
        root.destroy()
        if path:
            try:
                with open(path, "w", encoding="utf-8") as f:
                    f.write("RELATÓRIO DE MODIFICAÇÕES - SISTEMA DE TRANSMISSÕES\n")
                    f.write("="*60 + "\n\n")
                    for h in dados:
                        f.write(f"[{h.get('data_hora')}] {h.get('usuario')} - {h.get('acao')}\n")
                        f.write(f"DETALHES: {h.get('detalhes')}\n")
                        f.write("-" * 40 + "\n")
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Log exportado para TXT com sucesso!"), bgcolor=ft.Colors.GREEN_700)
            except Exception as e:
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Erro ao exportar TXT: {e}"), bgcolor=ft.Colors.RED_700)
            self.page.snack_bar.open = True; self.page.update()

    def abrir_opcoes_exportar(self):
        dlg = ft.AlertDialog(
            title=ft.Text("Escolha o formato de exportação"),
            content=ft.Column([
                ft.ListTile(
                    leading=ft.Icon(ft.Icons.TABLE_CHART, color=ft.Colors.GREEN_400),
                    title=ft.Text("Exportar para Excel (.xlsx)"),
                    subtitle=ft.Text("Ideal para edição e cálculos"),
                    on_click=lambda _: [self.fechar_dlg(dlg), self.exportar_excel()]
                ),
                ft.ListTile(
                    leading=ft.Icon(ft.Icons.TEXT_SNIPPET, color=ft.Colors.BLUE_400),
                    title=ft.Text("Exportar para TXT (.txt)"),
                    subtitle=ft.Text("Relatório em texto simples"),
                    on_click=lambda _: [self.fechar_dlg(dlg), self.exportar_txt()]
                ),
                ft.ListTile(
                    leading=ft.Icon(ft.Icons.DESCRIPTION, color=ft.Colors.ORANGE_400),
                    title=ft.Text("Exportar para CSV (.csv)"),
                    subtitle=ft.Text("Formato de dados universal"),
                    on_click=lambda _: [self.fechar_dlg(dlg), self.exportar_csv()]
                ),
            ], tight=True, spacing=10),
            actions=[ft.TextButton("Cancelar", on_click=lambda _: self.fechar_dlg(dlg))],
            shape=ft.RoundedRectangleBorder(radius=15),
        )
        self.page.overlay.append(dlg); dlg.open = True; self.page.update()

    def exportar_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from tkinter import filedialog
        import tkinter as tk

        # Filtra os dados atuais (reutiliza a lógica do exportar_csv)
        eventos_filtrados = self.get_eventos_filtrados()

        if not eventos_filtrados:
            self.page.snack_bar = ft.SnackBar(ft.Text("Nenhum dado para exportar!"), bgcolor=ft.Colors.ORANGE_800)
            self.page.snack_bar.open = True; self.page.update()
            return

        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile=f"historico_transmissoes_{datetime.now().strftime('%Y%m%d')}.xlsx")
        root.destroy()

        if path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Histórico"
                headers = ["Data", "Horário", "Evento", "Responsável", "Tipo", "Modalidade", "Local", "Operador", "Status"]
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                for t in eventos_filtrados:
                    ws.append([
                        format_date_br(t.data),
                        f"{t.horario_inicio} às {t.horario_fim}",
                        t.evento,
                        t.responsavel,
                        t.tipo_transmissao or "-",
                        t.modalidade or "-",
                        t.local or "-",
                        t.operador or "-",
                        t.status
                    ])
                
                # Ajuste de colunas
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try: max_length = max(max_length, len(str(cell.value)))
                        except: pass
                    ws.column_dimensions[column].width = min(max_length + 2, 50)

                wb.save(path)
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Exportado com sucesso: {os.path.basename(path)}"), bgcolor=ft.Colors.GREEN_700)
            except Exception as e:
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Erro ao exportar Excel: {e}"), bgcolor=ft.Colors.RED_700)
            self.page.snack_bar.open = True; self.page.update()

    def exportar_txt(self):
        from tkinter import filedialog
        import tkinter as tk
        eventos_filtrados = self.get_eventos_filtrados()
        if not eventos_filtrados: return
        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Arquivo de Texto", "*.txt")], initialfile=f"historico_transmissoes_{datetime.now().strftime('%Y%m%d')}.txt")
        root.destroy()
        if path:
            try:
                with open(path, "w", encoding="utf-8") as f:
                    f.write("HISTÓRICO DE TRANSMISSÕES\n")
                    f.write("="*60 + "\n\n")
                    for t in eventos_filtrados:
                        f.write(f"Evento: {t.evento}\n")
                        f.write(f"Data/Hora: {format_date_br(t.data)} às {t.horario_inicio}\n")
                        f.write(f"Responsável: {t.responsavel}\n")
                        f.write(f"Local: {t.local} | Status: {t.status}\n")
                        f.write("-" * 30 + "\n")
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Histórico exportado para TXT com sucesso!"), bgcolor=ft.Colors.GREEN_700)
            except Exception as e:
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Erro ao exportar TXT: {e}"), bgcolor=ft.Colors.RED_700)
            self.page.snack_bar.open = True; self.page.update()

    def get_eventos_filtrados(self):
        termo = self.txt_busca.value.lower() if self.txt_busca.value else ""
        status = self.dd_status.value
        tipo = self.dd_tipo.value
        modalidade = self.dd_modalidade.value
        local = self.dd_local.value
        operador = self.dd_operador.value
        periodo_tipo = self.dd_periodo_tipo.value
        mes_f = int(self.dd_filtro_mes.value) if self.dd_filtro_mes.value else None
        ano_f = int(self.dd_filtro_ano.value) if self.dd_filtro_ano.value else None
        
        hoje = datetime.now()
        hoje_str = hoje.strftime("%Y-%m-%d")
        
        dt_ini = None; dt_fim = None
        if periodo_tipo == "Personalizado":
            try:
                if self.txt_data_inicio.value: dt_ini = datetime.strptime(self.txt_data_inicio.value, "%d/%m/%Y")
                if self.txt_data_fim.value: dt_fim = datetime.strptime(self.txt_data_fim.value, "%d/%m/%Y")
            except: pass

        eventos_filtrados = []
        for t in self.controller.transmissoes:
            dt_t = parse_date(t.data)
            if not dt_t: continue
            dt_str = normalize_date(t.data)

            # 1. Filtro de Período
            if periodo_tipo == "Por Mês":
                if dt_t.month != mes_f or dt_t.year != ano_f: continue
            elif periodo_tipo == "Por Ano":
                if dt_t.year != ano_f: continue
            elif periodo_tipo == "Personalizado":
                if dt_ini and dt_t < dt_ini: continue
                if dt_fim and dt_t > dt_fim: continue
            else:
                if not termo and status == "Todos" and tipo == "Todos" and modalidade == "Todos" and local == "Todos" and operador == "Todos":
                    if dt_str > hoje_str: continue

            # 2. Busca por termo
            if termo:
                campos = [t.evento, t.responsavel, format_date_br(t.data), t.tipo_transmissao, t.modalidade, t.local, t.status, getattr(t, 'operador', '')]
                if not any(termo in str(c).lower() for c in campos if c): continue

            # 3. Dropdowns
            if not check_status_match(t.status, status): continue
            if tipo != "Todos" and t.tipo_transmissao != tipo: continue
            if modalidade != "Todos" and t.modalidade != modalidade: continue
            if local != "Todos" and t.local != local: continue
            if operador != "Todos" and getattr(t, 'operador', '') != operador: continue
            
            eventos_filtrados.append(t)
            
        return eventos_filtrados

    def update_year_menu(self):
        anos_disp = sorted(list(set(parse_date(t.data).year for t in self.controller.transmissoes if parse_date(t.data))), reverse=True)
        if not anos_disp: anos_disp = [datetime.now().year]
        self.menu_ano.items = [ft.PopupMenuItem(content=ft.Text(str(a)), on_click=lambda e, a=a: self.set_filtro_ano_bar(a)) for a in anos_disp]

    def set_filtro_mes_bar(self, mes, nome):
        self.dd_filtro_mes.value = str(mes) if mes > 0 else None
        self.txt_mes_selecionado.value = nome
        self.update_meses_bar()
        self.filtrar()

    def set_filtro_ano_bar(self, ano):
        self.dd_filtro_ano.value = str(ano)
        self.txt_ano_selecionado.value = str(ano)
        self.update_year_menu()
        self.update_meses_bar()
        self.filtrar()

    def update_meses_bar(self):
        self.row_meses_disponiveis.controls = []
        mes_atual_f = self.dd_filtro_mes.value
        
        # Abreviações dos meses
        meses_abrev = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
        
        # Filtra meses que possuem eventos no ano selecionado
        ano_selecionado = int(self.dd_filtro_ano.value) if self.dd_filtro_ano.value else datetime.now().year
        meses_com_eventos = set(parse_date(t.data).month for t in self.controller.transmissoes if parse_date(t.data) and parse_date(t.data).year == ano_selecionado)
        
        for i, nome in enumerate(meses_abrev):
            mes_num = i + 1
            is_selected = str(mes_num) == mes_atual_f
            has_events = mes_num in meses_com_eventos
            
            self.row_meses_disponiveis.controls.append(
                ft.TextButton(
                    content=ft.Text(
                        nome, 
                        size=11, 
                        weight=ft.FontWeight.BOLD if is_selected else ft.FontWeight.W_400,
                        color=ft.Colors.WHITE if is_selected else ft.Colors.WHITE38 if not has_events else ft.Colors.WHITE70
                    ),
                    style=ft.ButtonStyle(padding=0),
                    on_click=lambda e, m=mes_num, n=self.meses_nomes[i]: self.set_filtro_mes_bar(m, n)
                )
            )
        try: self.filter_bar.update()
        except: pass

    def toggle_selection(self, tid, checked):
        if checked: self.selected_ids.add(tid)
        else: self.selected_ids.discard(tid)
        self.btn_exportar_links.text = f"Gerar ({len(self.selected_ids)})" if self.selected_ids else "Exportar Links"
        self.update()

    def toggle_select_all(self, checked):
        eventos = self.get_eventos_filtrados()
        if checked:
            for t in eventos: self.selected_ids.add(t.id)
        else:
            for t in eventos: self.selected_ids.discard(t.id)
        self.btn_exportar_links.text = f"Gerar ({len(self.selected_ids)})" if self.selected_ids else "Exportar Links"
        self.filtrar()

    def cancelar_selecao_links(self, e):
        self.selection_mode = False
        self.selected_ids.clear()
        self.btn_exportar_links.text = "Exportar Links"
        self.btn_cancelar_links.visible = False
        self.selection_bar_control.visible = False
        self.tabela.columns[0].visible = False
        self.filtrar()

    def exportar_links(self, e):
        if not self.selection_mode:
            self.selection_mode = True
            self.btn_cancelar_links.visible = True
            self.tabela.columns[0].visible = True
            self.filtrar()
            return
        
        if not self.selected_ids:
            self.page.snack_bar = ft.SnackBar(ft.Text("Selecione pelo menos uma transmissão!"), bgcolor=ft.Colors.ORANGE_800)
            self.page.snack_bar.open = True; self.page.update()
            return
            
        from utils.dialog_helper import abrir_dialogo_links_multiplos
        transmissoes_selecionadas = [t for t in self.controller.transmissoes if t.id in self.selected_ids]
        abrir_dialogo_links_multiplos(self.page, transmissoes_selecionadas)

    def exportar_csv(self):
        import csv
        from tkinter import filedialog
        import tkinter as tk
        
        eventos_filtrados = self.get_eventos_filtrados()

        if not eventos_filtrados:
            self.page.snack_bar = ft.SnackBar(ft.Text("Nenhum dado para exportar!"), bgcolor=ft.Colors.ORANGE_800)
            self.page.snack_bar.open = True; self.page.update()
            return

        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")], initialfile=f"historico_transmissoes_{datetime.now().strftime('%Y%m%d')}.csv")
        root.destroy()

        if path:
            try:
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f, delimiter=";")
                    writer.writerow(["Data", "Horário", "Evento", "Responsável", "Tipo", "Modalidade", "Local", "Operador", "Status"])
                    for t in eventos_filtrados:
                        writer.writerow([
                            format_date_br(t.data),
                            f"{t.horario_inicio} às {t.horario_fim}",
                            t.evento,
                            t.responsavel,
                            t.tipo_transmissao or "-",
                            t.modalidade or "-",
                            t.local or "-",
                            t.operador or "-",
                            t.status
                        ])
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Exportado com sucesso: {os.path.basename(path)}"), bgcolor=ft.Colors.GREEN_700)
            except Exception as e:
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Erro ao exportar CSV: {e}"), bgcolor=ft.Colors.RED_700)
            self.page.snack_bar.open = True; self.page.update()
